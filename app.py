import streamlit as st
import openpyxl
import os
import tempfile
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak
import base64

def excel_to_pdf(input_excel_path, output_pdf_path):
    # Load the Excel file
    workbook = openpyxl.load_workbook(input_excel_path, data_only=True)

    # Create a PDF document
    doc = SimpleDocTemplate(output_pdf_path, pagesize=landscape(letter))

    # Iterate through each sheet and add a table to the PDF
    story = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        
        # Create a table with the data from the sheet
        table = Table(data)
        
        # Apply styles to the table (optional)
        table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                                   ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                                   ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                   ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                                   ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                                   ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                                   ('GRID', (0, 0), (-1, -1), 1, colors.black)]))
        
        # Add the table to the story and a page break
        story.append(table)
        story.append(PageBreak())

    # Build the PDF document
    doc.build(story)

def main():
    st.title("Excel to PDF Converter")

    uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx"])
    if uploaded_file is not None:
        st.write("File uploaded successfully!")

        convert_button = st.button("Convert to PDF")
        if convert_button:
            with st.spinner("Converting..."):
                with tempfile.NamedTemporaryFile(delete=False) as tmp_file:
                    output_pdf_path = tmp_file.name
                excel_to_pdf(uploaded_file, output_pdf_path)
                st.success("Conversion complete!")

            st.markdown(download_link_pdf(output_pdf_path), unsafe_allow_html=True)

def download_link_pdf(pdf_path):
    with open(pdf_path, "rb") as file:
        pdf_content = file.read()
    b64 = base64.b64encode(pdf_content).decode()
    return f'<a href="data:application/pdf;base64,{b64}" download="output.pdf">Click here to download the PDF</a>'

if __name__ == "__main__":
    main()
